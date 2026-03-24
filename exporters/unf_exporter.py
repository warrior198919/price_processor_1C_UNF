"""
Экспортёр в формат загрузки номенклатуры 1С УНФ (xlsx).
Создаёт файл с листом, совместимым с обработкой «Загрузка данных из файла» в 1С УНФ.

Формат колонок 1С УНФ:
  Наименование | Артикул | Единица измерения | Ставка НДС | Цена продажи |
  Описание | Группа номенклатуры | Бренд (производитель) |
  Штрихкод | Изображение (URL) | Полное наименование
"""

import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Колонки целевого файла 1С УНФ
UNF_COLUMNS = [
    "Наименование",
    "Артикул",
    "Единица измерения",
    "Ставка НДС",
    "Цена продажи (руб)",
    "Описание",
    "Группа номенклатуры",
    "Бренд (производитель)",
    "Штрихкод",
    "Изображение",
    "Полное наименование",
    "Vendorcode",
    "Подгруппа",
]


class UNFExporter:
    def __init__(self, records: list[dict], output_dir: str,
                 include_photos: bool = True,
                 split_by_group: bool = False,
                 price_field: str = "price_rub"):
        self.records = records
        self.output_dir = output_dir
        self.include_photos = include_photos
        self.split_by_group = split_by_group
        self.price_field = price_field  # "price_rub" or "price_ue"

    def export(self) -> list[str]:
        os.makedirs(self.output_dir, exist_ok=True)
        saved = []

        if self.split_by_group:
            groups = {}
            for r in self.records:
                g = r.get("Подгруппа") or "Прочее"
                groups.setdefault(g, []).append(r)
            for group, recs in groups.items():
                safe = _safe_filename(group)
                fname = f"1С_УНФ_{safe}_{_ts()}.xlsx"
                path = os.path.join(self.output_dir, fname)
                self._write_xlsx(recs, path)
                saved.append(path)
        else:
            brand = self.records[0].get("Бренд", "export") if self.records else "export"
            fname = f"1С_УНФ_{brand}_{_ts()}.xlsx"
            path = os.path.join(self.output_dir, fname)
            self._write_xlsx(self.records, path)
            saved.append(path)

        return saved

    def _write_xlsx(self, records: list[dict], path: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Номенклатура"

        # ── Заголовок ──────────────────────────────────────────────────────
        ws.append(UNF_COLUMNS)
        header_fill = PatternFill("solid", start_color="1A3A5C")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border

        ws.row_dimensions[1].height = 32

        # ── Данные ────────────────────────────────────────────────────────
        even_fill = PatternFill("solid", start_color="F2F6FB")
        data_font = Font(name="Arial", size=9)
        data_align = Alignment(vertical="center", wrap_text=False)

        for idx, rec in enumerate(records, start=2):
            price = rec.get("Цена розничная (руб)", 0) if self.price_field == "price_rub" \
                else rec.get("Цена розничная (у.е.)", 0)

            row_data = [
                rec.get("Номенклатура", ""),         # Наименование
                rec.get("Артикул", ""),               # Артикул
                rec.get("Ед. изм.", "шт."),           # Единица измерения
                "20%",                                 # Ставка НДС
                price,                                 # Цена продажи
                rec.get("Описание", ""),              # Описание
                rec.get("Full path", ""),             # Группа номенклатуры
                rec.get("Бренд", ""),                 # Бренд
                rec.get("Штрихкод", ""),              # Штрихкод
                rec.get("Фотографии", "") if self.include_photos else "",  # Изображение
                rec.get("Номенклатура", ""),          # Полное наименование
                rec.get("Vendorcode", ""),            # Vendorcode
                rec.get("Подгруппа", ""),             # Подгруппа
            ]

            ws.append(row_data)
            row_fill = even_fill if idx % 2 == 0 else None

            for col_idx, cell in enumerate(ws[idx], start=1):
                cell.font = data_font
                cell.alignment = data_align
                cell.border = border
                if row_fill:
                    cell.fill = row_fill
                # Форматирование цены
                if col_idx == 5 and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

        # ── Ширина колонок ────────────────────────────────────────────────
        col_widths = [55, 14, 14, 10, 16, 60, 40, 20, 16, 55, 55, 14, 30]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Заморозить первую строку
        ws.freeze_panes = "A2"

        # Автофильтр
        ws.auto_filter.ref = ws.dimensions

        # ── Лист «Инструкция» ─────────────────────────────────────────────
        info = wb.create_sheet("Инструкция")
        info.column_dimensions["A"].width = 80
        instructions = [
            ["Инструкция по загрузке в 1С УНФ"],
            [""],
            ["1. Откройте 1С УНФ → меню «Данные» → «Загрузка данных из файла»"],
            ["2. Укажите путь к этому файлу"],
            ["3. Выберите лист «Номенклатура»"],
            ["4. Сопоставьте колонки (сопоставление обычно происходит автоматически)"],
            ["5. Укажите режим: «Создавать новые» или «Обновлять существующие»"],
            ["6. Нажмите «Загрузить»"],
            [""],
            ["Колонки файла:"],
            ["  Наименование        — краткое название товара"],
            ["  Артикул             — артикул поставщика (Vendorcode)"],
            ["  Единица измерения   — шт., уп., кор. и т.д."],
            ["  Ставка НДС          — 20% (можно изменить перед загрузкой)"],
            ["  Цена продажи (руб)  — розничная цена в рублях"],
            ["  Описание            — полное описание товара"],
            ["  Группа номенклатуры — иерархический путь (Бренд\\Группа)"],
            ["  Бренд               — производитель"],
            ["  Штрихкод            — если есть"],
            ["  Изображение         — URL фотографии"],
            ["  Полное наименование — дублирует Наименование"],
            [""],
            [f"Файл создан: {datetime.now().strftime('%d.%m.%Y %H:%M')}"],
        ]
        for row in instructions:
            info.append(row)
        info["A1"].font = Font(bold=True, size=13, color="1A3A5C")

        wb.save(path)


# ── утилиты ───────────────────────────────────────────────────────────────

def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _safe_filename(s: str) -> str:
    import re
    s = re.sub(r'[\\/:*?"<>|]', "_", s)
    return s[:40].strip("_")
